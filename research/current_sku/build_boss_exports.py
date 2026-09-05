from __future__ import annotations
from pathlib import Path
import json, math, re
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(r"C:\Users\tarou\Downloads\sake-saijo-db")
WEB_DIR = ROOT / "research" / "web"
SKU_DIR = ROOT / "research" / "current_sku"
WEB = pd.read_csv(WEB_DIR / "index.csv", encoding="utf-8-sig")
SKU = pd.read_csv(SKU_DIR / "sku_master.csv", encoding="utf-8-sig")

AREAS = {
    "西条エリア7蔵": ["賀茂鶴酒造","白牡丹酒造","西條鶴醸造","亀齢酒造","福美人酒造","賀茂泉酒造","山陽鶴酒造"],
    "安芸津・黒瀬エリア3蔵": ["今田酒造本店","柄酒造","金光酒造"],
}

with (WEB_DIR / "sources.jsonl").open(encoding="utf-8") as f:
    SOURCE_MAP = {x["source_id"]: x for x in map(json.loads, f)}
EXCLUDE_TOKENS = [
    "酒まつり", "記念酒", "記念 ", "見学室直売所", "鑑評会出品酒", "オリジナルラベル",
    "飲み比べ", "セット", "アソート", "菰樽", "角樽", "JAL搭載限定ラベル",
    "蔵元限定", "酒蔵限定", "蔵内限定", "吉田屋の酒",
]
SEASONAL_TOKENS = ["しぼりたて", "ひやおろし", "秋上がり", "夏", "春", "新酒", "生酒", "うすにごり", "おりがらみ"]
LIMITED_TOKENS = ["流通限定", "取扱い店限定", "アメリカ限定", "海外限定", "輸出限定"]

ALIASES = {
    ("白牡丹酒造", "大吟醸 いちの雫 （しずく酒）"): "大吟醸 いちの雫（しずく酒）",
    ("賀茂泉酒造", "COKUN（こくん）"): "COKUN",
    ("賀茂泉酒造", "純米吟醸生原酒 RockHopper"): "純米吟醸生原酒 ROCK HOPPER",
    ("賀茂泉酒造", "朱泉普通酒紙パック"): "朱泉普通酒",
    ("賀茂泉酒造", "緑泉普通酒紙パック"): "緑泉普通酒",
    ("賀茂泉酒造", "造賀"): "造賀純米酒",
    ("賀茂鶴酒造", "本醸造 上等酒 カップ"): "本醸造 上等酒",
    ("賀茂鶴酒造", "純米酒 カップ"): "純米酒",
    ("亀齢酒造", "大吟醸 創"): "亀齢 大吟醸 創",
}

# COKUN+ は外部資料で法的分類が「その他の醸造酒」とされるため、日本酒（清酒）本表から除外。
SPECIAL_EXCLUDE = {("賀茂泉酒造", "COKUN+(こくんぷらす）"), ("賀茂泉酒造", "COKUN＋(プラス)")}
# 公開ウェブの追加調査で確認した飲用温度。__MAIN__ は主出典URLを使用。
SUPPLEMENTAL_DRINK = {
    ("亀齢酒造","亀齢 純米吟醸 六拾"): ("冷酒（提供例）","現行飲食店での提供温度","https://benbefukuyamaekimae.owst.jp/drinks"),
    ("山陽鶴酒造","大吟醸"): ("冷酒","現行地域商品紹介","https://japan-stroll.com/sightseeing/nontasakagura/"),
    ("山陽鶴酒造","上撰"): ("常温 / 熱燗（提供例）","現行飲食店での提供温度","https://www.hotpepper.jp/strJ000661918/drink/"),
    ("白牡丹酒造","大吟醸 いちの雫（しずく酒）"): ("常温 / 冷酒","公式商品ページ","__MAIN__"),
    ("白牡丹酒造","長期低温熟成 大吟醸 FLUSSO"): ("常温 / 冷酒","公式商品ページ","__MAIN__"),
    ("白牡丹酒造","純米生酒 氷華"): ("常温 / 冷酒","公式商品ページ","__MAIN__"),
    ("白牡丹酒造","山田錦 純米大吟醸 氷華"): ("常温 / 冷酒","公式商品ページ","__MAIN__"),
    ("白牡丹酒造","広島八反 純米大吟醸 氷華"): ("常温 / 冷酒","公式商品ページ","__MAIN__"),
    ("西條鶴醸造","吟醸酒「ゴールド西條鶴」"): ("冷酒 / 常温 / ぬる燗","現行酒販店商品情報","https://osake-style.com/SHOP/10020122.html"),
    ("西條鶴醸造","純米生原酒"): ("冷酒","地域観光資料","https://higashihiroshima-digital-sightseeing.com/joy/ichioshi/"),
    ("西條鶴醸造","純米酒「大地の風」"): ("冷酒 / 常温 / 熱燗","地域観光資料","https://higashihiroshima-digital-sightseeing.com/joy/ichioshi/"),
    ("西條鶴醸造","純米大吟醸原酒「神髄」"): ("常温","広島県資料","https://www.pref.hiroshima.lg.jp/uploaded/attachment/524371.pdf"),
    ("西條鶴醸造","純米大吟醸「西鶴」"): ("常温","広島県資料","https://www.pref.hiroshima.lg.jp/uploaded/attachment/524371.pdf"),
    ("亀齢酒造","純米大吟醸 亀香"): ("常温 / 冷酒","地域観光資料","https://higashihiroshima-digital-sightseeing.com/joy/ichioshi/"),
    ("福美人酒造","特別純米酒「ひめあま」"): ("冷酒","酒蔵紹介資料","https://www.goshu-pro.jp/chugokushikoku/hiroshima/2415"),
    ("福美人酒造","大吟醸 蔵乃華 （ くらのはな ）"): ("常温 / 冷酒","酒造組合系商品資料","https://www.japansake.or.jp/sake/about/kumiaiin/pdf/34.pdf"),
    ("福美人酒造","大吟醸 西條酒造学校"): ("冷酒 / 常温","現行酒販店商品情報","https://osake-style.com/SHOP/10020036.html"),
    ("賀茂泉酒造","造賀純米酒"): ("常温 / ぬる燗","地域観光資料","https://higashihiroshima-digital-sightseeing.com/joy/ichioshi/"),
    ("山陽鶴酒造","本醸造"): ("冷酒 / 常温 / 燗酒","地域観光資料","https://higashihiroshima-digital-sightseeing.com/joy/ichioshi/"),
    ("柄酒造","9代目於多福 protos. 火入れ"): ("冷酒 / 常温","現行酒販店商品情報","https://tawawasake.base.shop/items/134292512"),
    ("金光酒造","賀茂金秀 純米大吟醸35"): ("冷酒（10～15℃前後）","日本酒情報DB","https://www.sakenomy.jp/sake/TST0000041706/"),
    ("金光酒造","賀茂金秀 純米大吟醸40"): ("冷酒（10～15℃前後）","日本酒情報DB","https://www.sakenomy.jp/sake/TST0000007476/"),
    ("金光酒造","賀茂金秀 純米吟醸 雄町"): ("冷酒（10～15℃前後）","日本酒情報DB","https://www.sakenomy.jp/sake/TST0000007478/"),
    ("金光酒造","賀茂金秀 特別純米"): ("冷酒 / 燗酒（45～50℃）","現行酒販店商品情報","https://reika-sake.com/products/34kmk01201121"),
    ("金光酒造","賀茂金秀 特別純米13"): ("冷酒（5～15℃前後）","現行酒販店商品情報","https://reika-sake.com/products/34kmk01203121"),
    ("金光酒造","賀茂金秀 辛口特別純米"): ("冷酒（10～15℃前後）","日本酒情報DB","https://www.sakenomy.jp/sake/TST0000010137/"),
}

SUPPLEMENTAL_METER = {
    ("金光酒造","賀茂金秀 純米大吟醸35"): "+1",
    ("金光酒造","賀茂金秀 純米大吟醸40"): "+1",
    ("金光酒造","賀茂金秀 純米吟醸 雄町"): "+2",
    ("金光酒造","賀茂金秀 特別純米"): "+3",
    ("金光酒造","賀茂金秀 特別純米13"): "-1",
    ("金光酒造","賀茂金秀 辛口特別純米"): "+8",
}

# WEB商品層に未収録でも、公式ECで現在販売中の主要商品は商品一覧へ補完する。
# 網羅目的で全SKUを昇格させず、代表性が高く独立商品として扱うべきものだけを明示管理する。
CURRENT_SKU_PRODUCT_SPECS = [
    {"酒蔵":"賀茂鶴酒造","商品名":"純米吟醸 一滴入魂","match":"純米吟醸 一滴入魂","分類":"純米吟醸","流通区分":"一般流通","出典URL":"https://shop.kamotsuru.jp/SHOP/itteki720.html","include_soldout":True},
    {"酒蔵":"賀茂鶴酒造","商品名":"賀茂鶴 光壽","match":"賀茂鶴光壽","分類":"その他","精米歩合":"28%","流通区分":"一般流通","出典URL":"https://shop.kamotsuru.jp/SHOP/kouju750.html"},
    {"酒蔵":"賀茂鶴酒造","商品名":"純米大吟醸 瑞兆賀茂鶴","match":"純米大吟醸 瑞兆賀茂鶴","分類":"純米大吟醸","流通区分":"一般流通","出典URL":"https://shop.kamotsuru.jp/SHOP/zuicho720.html"},
    {"酒蔵":"賀茂鶴酒造","商品名":"大吟醸 吉祥 賀茂鶴","match":"大吟醸 吉祥 賀茂鶴","分類":"大吟醸","流通区分":"一般流通","出典URL":"https://shop.kamotsuru.jp/SHOP/kissho720.html"},
    {"酒蔵":"賀茂鶴酒造","商品名":"大吟醸 天凜","match":"大吟醸 天凜","分類":"大吟醸","流通区分":"一般流通","出典URL":"https://shop.kamotsuru.jp/SHOP/tenrin720.html"},
    {"酒蔵":"賀茂鶴酒造","商品名":"大吟醸 吟凛雅","match":"大吟醸 吟凛雅","分類":"大吟醸","流通区分":"一般流通","出典URL":"https://shop.kamotsuru.jp/SHOP/ginringa900.html"},
    {"酒蔵":"賀茂鶴酒造","商品名":"酒中在心 鶯 純米大吟醸 山田錦","match":"酒中在心 鶯 純米大吟醸 山田錦","分類":"純米大吟醸","流通区分":"限定流通","出典URL":"https://shop.kamotsuru.jp/SHOP/shutyuzaishin-uguisu_720.html"},
]

def clean(v):
    if pd.isna(v): return ""
    return " ".join(str(v).replace("　", " ").split())
def canonical_name(brewery: str, name: str) -> str:
    name = clean(name)
    if (brewery, name) in ALIASES:
        return ALIASES[(brewery, name)]
    if brewery == "賀茂泉酒造" and name.startswith("COKUN+"):
        return "COKUN+"
    return name

def norm_name(name: str) -> str:
    s = clean(name).lower()
    s = re.sub(r"[【】\[\]（）()「」『』〈〉<>・･\s]", "", s)
    s = re.sub(r"\d+(?:\.\d+)?\s*(?:ml|mℓ|ｍｌ|l|ℓ)", "", s, flags=re.I)
    s = re.sub(r"(?:化粧箱入り?|木箱入り?|箱入|瓶詰|詰|カップ|紙パック)$", "", s)
    for p in ["賀茂鶴", "白牡丹", "西條鶴", "亀齢", "福美人", "賀茂泉", "富久長"]:
        if s.startswith(p.lower()): s = s[len(p):]
    return s

def parse_meter(v):
    s = clean(v).replace("＋", "+").replace("−", "-").replace("～", "~")
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else None

def sweetness(name: str, meter):
    if "辛口" in name:
        return "辛口", "商品名に辛口表記"
    if "甘口" in name or "甘い口どけ" in name:
        return "甘口", "商品名・商品説明上の甘口表記"
    x = parse_meter(meter)
    if x is None: return "不明", "公開情報で判定材料不足"
    if x <= -4: return "甘口", "日本酒度からの目安"
    if x < -1: return "やや甘口", "日本酒度からの目安"
    if x <= 1: return "中口", "日本酒度からの目安"
    if x < 4: return "やや辛口", "日本酒度からの目安"
    return "辛口", "日本酒度からの目安"

def category_bucket(designation: str, name: str) -> str:
    t = clean(designation) + " " + name
    if "純米大吟醸" in t: return "純米大吟醸"
    if "大吟醸" in t: return "大吟醸"
    if "純米吟醸" in t: return "純米吟醸"
    if "吟醸" in t: return "吟醸"
    if "特別純米" in t: return "特別純米"
    if "純米" in t: return "純米"
    if "特別本醸造" in t: return "特別本醸造"
    if "本醸造" in t: return "本醸造"
    if "普通酒" in t or "上撰" in t or "金紋" in t: return "普通酒・上撰"
    return clean(designation) or "その他"

def source_url(row):
    sid = clean(row.get("主出典ID"))
    return SOURCE_MAP.get(sid, {}).get("url", "")

def is_excluded(brewery: str, name: str):
    if (brewery, name) in SPECIAL_EXCLUDE or canonical_name(brewery, name) == "COKUN+":
        return True, "清酒以外の可能性があるため本表対象外"
    hit = next((x for x in EXCLUDE_TOKENS if x in name), "")
    if hit: return True, f"一般流通商品ではない／資料用途に不向き（{hit}）"
    if re.search(r"(?:202[0-5]年|令和[1-7]年|202[0-5])", name):
        return True, "過年度・年次限定商品として一般流通本表から除外"
    return False, ""
def sku_matches(brewery: str, product_name: str):
    cand = SKU[SKU["酒蔵"] == brewery].copy()
    pn = norm_name(product_name)
    hits = []
    for _, r in cand.iterrows():
        sn = norm_name(clean(r["商品名"]))
        if not pn or not sn: continue
        ratio = SequenceMatcher(None, pn, sn).ratio()
        if brewery == "賀茂鶴酒造" and pn == norm_name("純米吟醸") and "一滴入魂" in clean(r["商品名"]):
            continue
        contain = (len(pn) >= 4 and (pn in sn or sn in pn))
        if contain: ratio=max(ratio,0.95)
        if contain or ratio >= 0.76:
            hits.append((ratio, r))
    hits.sort(key=lambda x: x[0], reverse=True)
    return [r for _, r in hits if _ >= (hits[0][0] - 0.12 if hits else 1)]

def aggregate_sku(brewery: str, product_name: str):
    hits = sku_matches(brewery, product_name)
    caps, prices, channels = [], [], []
    for r in hits:
        v = r.get("容量ml")
        try:
            if pd.notna(v) and 100 <= float(v) <= 5000: caps.append(int(float(v)))
        except Exception: pass
        p = r.get("価格円税込")
        sku_name=clean(r.get("商品名"))
        is_multi=bool(re.search(r"(?:[2-9]|1[0-9]|2[0-9]|30)本|セット|アソート|ワイングラス",sku_name))
        try:
            if pd.notna(p) and float(p) > 0 and not is_multi: prices.append(int(float(p)))
        except Exception: pass
        ch = clean(r.get("販売チャネル"))
        if ch and ch not in channels: channels.append(ch)
    cap_txt = " / ".join(f"{x:,}ml" for x in sorted(set(caps)))
    price_txt = ""
    if prices:
        lo, hi = min(prices), max(prices)
        price_txt = f"{lo:,}円" if lo == hi else f"{lo:,}～{hi:,}円"
    return cap_txt, price_txt, " / ".join(channels)

def current_sku_product_row(spec):
    all_hits = SKU[SKU["酒蔵"] == spec["酒蔵"]].copy()
    all_hits = all_hits[all_hits["商品名"].fillna("").str.contains(spec["match"], regex=False)]
    all_hits = all_hits[all_hits["出典種別"].fillna("").str.startswith("official")]
    selling = all_hits[all_hits["販売状態"] == "販売中"]
    if selling.empty:
        return None
    hits = selling.copy()
    if spec.get("include_soldout"):
        soldout = all_hits[all_hits["在庫状態"] == "売切"]
        hits = pd.concat([hits, soldout], ignore_index=True).drop_duplicates(subset=["URL", "容量ml", "商品名"])
    caps=[]; prices=[]
    for _, sr in hits.iterrows():
        try:
            v=float(sr.get("容量ml"));
            if 100 <= v <= 5000: caps.append(int(v))
        except Exception: pass
        try:
            p=float(sr.get("価格円税込"));
            if p > 0: prices.append(int(p))
        except Exception: pass
    cap_txt=" / ".join(f"{x:,}ml" for x in sorted(set(caps)))
    price_txt=""
    if prices:
        lo,hi=min(prices),max(prices); price_txt=f"{lo:,}円" if lo==hi else f"{lo:,}～{hi:,}円"
    return {"酒蔵":spec["酒蔵"],"商品名":spec["商品名"],"分類":spec["分類"],"甘辛目安":"不明",
            "甘辛判定根拠":"公開情報で判定材料不足","日本酒度":"","精米歩合":spec.get("精米歩合",""),
            "アルコール度数":"","おすすめの飲み方":"要確認","飲み方情報":"公開情報で要確認",
            "飲み方出典URL":"","使用米":"","容量":cap_txt,"参考価格":price_txt,"流通区分":spec["流通区分"],
            "備考":"公式直営ECで販売中を確認。SKU層から商品層へ補完。","出典URL":spec["出典URL"],
            "出典数":int(hits["URL"].nunique())}

def build_products(breweries):
    rows, audit = [], []
    current = WEB[(WEB["酒蔵"].isin(breweries)) & (WEB["販売状態"] == "販売中")].copy()
    current["canon"] = [canonical_name(b,n) for b,n in zip(current["酒蔵"], current["商品名"])]
    for (brewery, canon), g in current.groupby(["酒蔵","canon"], sort=False):
        excluded, reason = is_excluded(brewery, canon)
        if excluded:
            for _, r in g.iterrows(): audit.append([brewery, clean(r["商品名"]), "除外", reason, ""])
            continue
        # 同一商品の複数表記は、出典数・項目充足度が高い行を代表として採用
        scored = []
        for _, r in g.iterrows():
            filled = sum(bool(clean(r.get(c))) for c in ["特定名称","精米歩合","アルコール度数","日本酒度","酸度","おすすめ温度","使用米"])
            scored.append((int(r.get("出典数") or 0) * 10 + filled, r))
        scored.sort(key=lambda x: x[0], reverse=True)
        r = scored[0][1]
        if len(g) > 1:
            for _, rr in g.iterrows():
                if clean(rr["商品名"]) != clean(r["商品名"]):
                    audit.append([brewery, clean(rr["商品名"]), "統合", "表記・包装違いを商品単位に統合", canon])
        main_url = source_url(r)
        meter = clean(r.get("日本酒度")) or SUPPLEMENTAL_METER.get((brewery,canon), "")
        sweet, sweet_basis = sweetness(canon, meter)
        drink = clean(r.get("おすすめ温度"))
        drink_info = "出典記載" if drink else "公開情報で要確認"
        drink_url = main_url if drink else ""
        if not drink and (brewery,canon) in SUPPLEMENTAL_DRINK:
            drink, drink_info, drink_url = SUPPLEMENTAL_DRINK[(brewery,canon)]
            if drink_url == "__MAIN__": drink_url = main_url
        cap, price, channel = aggregate_sku(brewery, canon)
        if not cap: cap = clean(r.get("容量"))
        if not price: price = clean(r.get("価格"))
        limited = ""
        if any(t in canon for t in LIMITED_TOKENS) or "取扱店限定" in channel or "流通限定" in channel: limited = "限定流通"
        elif any(t in canon for t in SEASONAL_TOKENS): limited = "季節商品"
        elif "輸出" in channel or "海外" in channel: limited = "輸出・海外向け"
        else: limited = "一般流通"
        row = {
            "酒蔵": brewery,
            "商品名": canon,
            "分類": category_bucket(r.get("特定名称"), canon),
            "甘辛目安": sweet,
            "甘辛判定根拠": sweet_basis,
            "日本酒度": meter,
            "精米歩合": clean(r.get("精米歩合")),
            "アルコール度数": clean(r.get("アルコール度数")),
            "おすすめの飲み方": drink or "要確認",
            "飲み方情報": drink_info,
            "飲み方出典URL": drink_url,
            "使用米": clean(r.get("使用米")),
            "容量": cap,
            "参考価格": price,
            "流通区分": limited,
            "備考": clean(r.get("備考")),
            "出典URL": main_url,
            "出典数": int(r.get("出典数") or 0),
        }
        rows.append(row)
    existing={(r["酒蔵"], r["商品名"]) for r in rows}
    for spec in CURRENT_SKU_PRODUCT_SPECS:
        if spec["酒蔵"] not in breweries or (spec["酒蔵"], spec["商品名"]) in existing:
            continue
        row=current_sku_product_row(spec)
        if row is not None:
            rows.append(row); existing.add((spec["酒蔵"],spec["商品名"]))
            audit.append([spec["酒蔵"], spec["match"], "SKU補完", "公式ECで販売中の主要独立商品を商品層へ補完", spec["商品名"]])
    df = pd.DataFrame(rows)
    if "亀齢酒造" in breweries and (df.empty or not ((df["酒蔵"]=="亀齢酒造") & (df["商品名"]=="亀齢 上撰")).any()):
        kirei_j={"酒蔵":"亀齢酒造","商品名":"亀齢 上撰","分類":"本醸造","甘辛目安":"辛口","甘辛判定根拠":"現行販売店の商品説明","日本酒度":"","精米歩合":"","アルコール度数":"","おすすめの飲み方":"冷酒 / 燗酒","飲み方情報":"現行販売店の商品説明","飲み方出典URL":"https://osake-style.com/SHOP/11012948.html","使用米":"","容量":"1,800ml","参考価格":"2,398円","流通区分":"一般流通","備考":"現行酒販店で在庫ありを確認","出典URL":"https://osake-style.com/SHOP/11012948.html","出典数":2}
        df=pd.concat([df,pd.DataFrame([kirei_j])],ignore_index=True)
    if "亀齢酒造" in breweries and (df.empty or not ((df["酒蔵"]=="亀齢酒造") & (df["商品名"]=="亀齢 純米吟醸 六拾")).any()):
        kirei60={"酒蔵":"亀齢酒造","商品名":"亀齢 純米吟醸 六拾","分類":"純米吟醸","甘辛目安":"不明","甘辛判定根拠":"公開情報で判定材料不足","日本酒度":"","精米歩合":"","アルコール度数":"","おすすめの飲み方":"要確認","飲み方情報":"公開情報で要確認","飲み方出典URL":"","使用米":"","容量":"720ml","参考価格":"","流通区分":"一般流通","備考":"2026年5月 東広島10蔵定期便採用","出典URL":"https://e-sake.jp/article/article-1369/","出典数":1}
        df=pd.concat([df,pd.DataFrame([kirei60])],ignore_index=True)
    if "山陽鶴酒造" in breweries and (df.empty or not ((df["酒蔵"]=="山陽鶴酒造") & (df["商品名"]=="純米大吟醸 KUBO")).any()):
        kubo={
            "酒蔵":"山陽鶴酒造","商品名":"純米大吟醸 KUBO","分類":"純米大吟醸","甘辛目安":"辛口","甘辛判定根拠":"日本酒度からの目安",
            "日本酒度":"+4","精米歩合":"35%","アルコール度数":"16度以上17度以下","おすすめの飲み方":"冷酒","飲み方情報":"現行返礼品で確認","飲み方出典URL":"https://www.furusato.aeon.co.jp/gift-in-return/aaa72316281ae6d75280fb28cb3cae46/",
            "使用米":"山田錦100%","容量":"720ml","参考価格":"","流通区分":"一般流通","備考":"2026年6月 東広島10蔵定期便採用／通年返礼品で提供",
            "出典URL":"https://www.furusato.aeon.co.jp/gift-in-return/aaa72316281ae6d75280fb28cb3cae46/","出典数":2,
        }
        df=pd.concat([df,pd.DataFrame([kubo])],ignore_index=True)
    if not df.empty:
        order = {b:i for i,b in enumerate(breweries)}
        df["_order"] = df["酒蔵"].map(order)
        df = df.sort_values(["_order","分類","商品名"]).drop(columns="_order").reset_index(drop=True)
    base_audit=pd.DataFrame(audit, columns=["酒蔵","元データ","処理","理由","統合先"]); sku_audit=build_sku_audit(df,breweries); return df, pd.concat([base_audit,sku_audit],ignore_index=True).drop_duplicates().reset_index(drop=True)
def build_sku_audit(df, breweries):
    out=[]
    for _,sr in SKU[SKU["酒蔵"].isin(breweries)].iterrows():
        b=clean(sr["酒蔵"]); n=clean(sr["商品名"])
        if any(t in n for t in ["セット","アソート","ギフト","飲み比べ","オリジナルラベル"]):
            out.append([b,n,"除外","セット・ギフト・特別包装SKUのため商品一覧から除外",""]); continue
        candidates=df[df["酒蔵"]==b]
        sn=norm_name(n); best=None; best_score=0
        for _,pr in candidates.iterrows():
            pn=norm_name(pr["商品名"])
            if not sn or not pn: continue
            score=SequenceMatcher(None,sn,pn).ratio()
            if len(pn)>=4 and (pn in sn or sn in pn): score=max(score,0.88)
            if score>best_score: best_score=score; best=pr["商品名"]
        if best is not None and best_score>=0.72:
            out.append([b,n,"SKU統合","容量・包装・商品コード違いを商品単位へ統合",best])
        else:
            out.append([b,n,"未採用","一般流通の商品単位へ確実に対応づけられないため本表には不採用",""])
    return pd.DataFrame(out,columns=["酒蔵","元データ","処理","理由","統合先"])

def main_score(r):
    name = r["商品名"]
    score = min(int(r["出典数"]), 4) * 2
    score += 2 if r["分類"] != "その他" else 0
    score += 2 if r["おすすめの飲み方"] != "要確認" else 0
    score += 1 if r["日本酒度"] else 0
    score += 1 if r["精米歩合"] else 0
    if any(t in name for t in SEASONAL_TOKENS): score -= 5
    if any(t in name for t in ["限定", "原酒", "にごり", "おりがらみ"]): score -= 3
    if any(t in name for t in ["カップ", "パック"]): score -= 4
    return score

def choose_main(df, breweries):
    out=[]
    priority=["純米大吟醸","大吟醸","純米吟醸","吟醸","特別純米","純米","特別本醸造","本醸造","普通酒・上撰","その他"]
    for brewery in breweries:
        g=df[df["酒蔵"]==brewery].copy()
        if g.empty: continue
        g["_score"]=g.apply(main_score,axis=1)
        stable=g[(g["流通区分"]=="一般流通") & ~g["商品名"].apply(lambda x:any(t in x for t in SEASONAL_TOKENS+["限定"]))].copy()
        pool=stable if len(stable)>=5 else g
        chosen=[]; names=set()
        for cat in priority:
            cg=pool[pool["分類"]==cat].sort_values(["_score","出典数"],ascending=False)
            if not cg.empty:
                r=cg.iloc[0]
                if r["商品名"] not in names: chosen.append(r); names.add(r["商品名"])
                if len(chosen)==5: break
        if len(chosen)<5:
            for _,r in pool.sort_values(["_score","出典数"],ascending=False).iterrows():
                if r["商品名"] in names: continue
                chosen.append(r); names.add(r["商品名"])
                if len(chosen)==5: break
        for r in chosen:
            d=r.to_dict(); d["選定理由"]=f"現行流通・{d['分類']}枠／定番性と酒質の幅を考慮"; d.pop("_score",None); out.append(d)
    cols=["酒蔵","商品名","分類","甘辛目安","日本酒度","おすすめの飲み方","容量","流通区分","選定理由","出典URL"]
    return pd.DataFrame(out)[cols] if out else pd.DataFrame(columns=cols)

NAVY = "17365D"; BLUE = "2F75B5"; LIGHT = "D9EAF7"; PALE = "F3F6F9"; WHITE = "FFFFFF"; GRAY = "666666"
THIN = Side(style="thin", color="D9E1F2")

def setup_sheet(ws, title, subtitle, headers, rows):
    ws.sheet_view.showGridLines = False
    end_col = max(1, len(headers))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1,1,title); c.font=Font(size=16,bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=end_col)
    c=ws.cell(2,1,subtitle); c.font=Font(size=9,color=GRAY); c.alignment=Alignment(wrap_text=True,vertical="center")
    ws.row_dimensions[2].height=30
    for j,h in enumerate(headers,1):
        cell=ws.cell(4,j,h); cell.font=Font(bold=True,color=WHITE); cell.fill=PatternFill("solid",fgColor=BLUE); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=Border(bottom=THIN)
    for i,row in enumerate(rows,5):
        for j,v in enumerate(row,1):
            cell=ws.cell(i,j,"" if v is None or (isinstance(v,float) and math.isnan(v)) else v)
            cell.alignment=Alignment(vertical="top",wrap_text=True)
            cell.border=Border(bottom=THIN)
            if i%2==0: cell.fill=PatternFill("solid",fgColor=PALE)
    if rows:
        ref=f"A4:{get_column_letter(end_col)}{4+len(rows)}"
        tab=Table(displayName=f"T_{abs(hash(title))%1000000}", ref=ref)
        tab.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showFirstColumn=False,showLastColumn=False,showRowStripes=False,showColumnStripes=False)
        ws.add_table(tab)
    ws.freeze_panes="A5"
    ws.auto_filter.ref=f"A4:{get_column_letter(end_col)}{4+max(1,len(rows))}"

def widths(ws, mapping):
    for col,w in mapping.items(): ws.column_dimensions[col].width=w

def add_url_links(ws, header_row=4):
    headers={ws.cell(header_row,c).value:c for c in range(1,ws.max_column+1)}
    for header in ["出典URL","飲み方出典URL"]:
        col=headers.get(header)
        if not col: continue
        for r in range(header_row+1,ws.max_row+1):
            cell=ws.cell(r,col)
            if isinstance(cell.value,str) and cell.value.startswith("http"):
                cell.hyperlink=cell.value; cell.style="Hyperlink"

def build_summary_sheet(wb, area, df, breweries):
    ws=wb.active; ws.title="概要"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:H1"); ws["A1"]=f"東広島市 {area} 日本酒リスト"; ws["A1"].font=Font(size=18,bold=True,color=WHITE); ws["A1"].fill=PatternFill("solid",fgColor=NAVY); ws["A1"].alignment=Alignment(vertical="center"); ws.row_dimensions[1].height=32
    lines=[
        "対象：現在販売中と確認できる一般流通の独立商品。容量・包装違いは商品単位に統合。",
        "除外：酒まつり限定、記念酒、鑑評会出品酒、直売所記念品、ギフトセット等。流通限定・取扱店限定は市場流通品として残しています。",
        "甘辛目安：蔵元等の明示がない場合、日本酒度を参考に便宜上分類。酸度等により実際の味覚は異なるため『目安』です。",
        "おすすめの飲み方：公開出典の推奨を優先。一部は現行飲食店の提供温度を『提供例』と明記。根拠がないものは『要確認』です。",
    ]
    for i,t in enumerate(lines,3):
        ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=8); ws.cell(i,1,t).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(i,1).font=Font(size=10)
    ws["A8"]="酒蔵"; ws["B8"]="掲載商品数"; ws["C8"]="うち飲み方確認済"
    for c in ws[8]: c.font=Font(bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=BLUE); c.alignment=Alignment(horizontal="center")
    for i,b in enumerate(breweries,9):
        g=df[df["酒蔵"]==b]; ws.cell(i,1,b); ws.cell(i,2,len(g)); ws.cell(i,3,int((g["おすすめの飲み方"]!="要確認").sum()))
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=16; ws.column_dimensions["C"].width=20
    for c in range(4,9): ws.column_dimensions[get_column_letter(c)].width=12

def write_main_book(path, area, df, main, breweries):
    wb=Workbook(); build_summary_sheet(wb,area,df,breweries)
    ws=wb.create_sheet("主要銘柄")
    headers=list(main.columns); rows=main.fillna("").values.tolist()
    setup_sheet(ws,f"東広島市 {area} 主要日本酒一覧", "各酒蔵 最大5商品。人気順位ではなく、現行流通・酒質の幅・公開情報の充実度を基準に選定。",headers,rows)
    widths(ws,{"A":18,"B":34,"C":16,"D":12,"E":12,"F":28,"G":18,"H":16,"I":38,"J":42}); add_url_links(ws)
    crit=wb.create_sheet("選定基準"); setup_sheet(crit,"主要銘柄の選定基準","上司向けの俯瞰表として、同系統の商品だけに偏らないよう整理しています。",["基準","内容"],[
        ["対象","現時点で販売中と確認できる一般流通商品"],
        ["代表性","出典数や公式掲載状況を参考に、酒蔵を説明しやすい商品を優先"],
        ["多様性","純米大吟醸・吟醸・純米・本醸造など分類が重複しすぎないよう選定"],
        ["安定性","イベント限定・記念酒・直売所限定・季節商品は主要5選では優先しない"],
        ["注意","本表は売上・人気ランキングではありません"],
    ]); widths(crit,{"A":18,"B":80})
    wb.save(path)
def write_full_book(path, area, df, audit, breweries):
    wb=Workbook(); build_summary_sheet(wb,area,df,breweries)
    ws=wb.create_sheet("現行商品一覧")
    show_cols=["酒蔵","商品名","分類","甘辛目安","甘辛判定根拠","日本酒度","精米歩合","アルコール度数","おすすめの飲み方","飲み方情報","飲み方出典URL","使用米","容量","参考価格","流通区分","備考","出典URL"]
    data=df[show_cols].copy(); setup_sheet(ws,f"東広島市 {area} 現行日本酒一覧","容量・包装違いは統合済み。生のSKU一覧ではなく、上司説明用の商品単位です。",show_cols,data.fillna("").values.tolist())
    widths(ws,{"A":18,"B":36,"C":16,"D":12,"E":22,"F":12,"G":14,"H":16,"I":28,"J":18,"K":20,"L":20,"M":18,"N":18,"O":36,"P":44}); add_url_links(ws)

    agg=wb.create_sheet("分類別集計")
    pivot=pd.crosstab(df["酒蔵"],df["分類"]).reindex(breweries,fill_value=0)
    pheaders=["酒蔵"]+pivot.columns.tolist()+["合計"]
    prows=[]
    for b,row in pivot.iterrows(): prows.append([b]+[int(x) for x in row.tolist()]+[int(row.sum())])
    setup_sheet(agg,f"東広島市 {area} 分類別集計","商品単位に丸めた後の件数です。",pheaders,prows)
    for c in range(1,len(pheaders)+1): agg.column_dimensions[get_column_letter(c)].width=16

    aud=wb.create_sheet("除外・統合一覧")
    aheaders=list(audit.columns); arows=audit.fillna("").values.tolist()
    setup_sheet(aud,"除外・統合一覧","SKU・包装・イベント品などを、上司向け本表へ載せない／統合した理由を残しています。",aheaders,arows)
    widths(aud,{"A":18,"B":42,"C":12,"D":52,"E":36})
    wb.save(path)

def main():
    results={}
    for area, breweries in AREAS.items():
        df,audit=build_products(breweries)
        main_df=choose_main(df,breweries)
        prefix=f"東広島市_{area}"
        main_path=ROOT / f"{prefix}_主要日本酒一覧.xlsx"
        full_path=ROOT / f"{prefix}_現行日本酒一覧.xlsx"
        write_main_book(main_path,area,df,main_df,breweries)
        write_full_book(full_path,area,df,audit,breweries)
        results[area]={"products":len(df),"main":len(main_df),"audit":len(audit),"main_file":main_path.name,"full_file":full_path.name,"by_brewery":df["酒蔵"].value_counts().reindex(breweries,fill_value=0).to_dict()}
        df.to_csv(SKU_DIR/f"boss_{area}_products.csv",index=False,encoding="utf-8-sig")
        audit.to_csv(SKU_DIR/f"boss_{area}_audit.csv",index=False,encoding="utf-8-sig")
    (SKU_DIR/"boss_export_summary.json").write_text(json.dumps(results,ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps(results,ensure_ascii=False,indent=2))

if __name__ == "__main__": main()
